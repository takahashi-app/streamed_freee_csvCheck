"""
Excel出力ユーティリティ
DataFrameをExcelに出力し、色分けを行う
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelWriter:
    """Excel出力を行うクラス"""

    # 色定義
    COLOR_GREEN = 'C6EFCE'  # 完全一致（緑）
    COLOR_RED = 'FFC7CE'    # 候補なし・低スコア（赤）
    COLOR_YELLOW = 'FFEB9C'  # 取引先候補1（黄色）
    COLOR_YELLOW_LIGHT = 'FFF9E6'  # 取引先元・候補2-3（薄い黄色）
    COLOR_BLUE = 'DDEBF7'    # 部門候補1（青色）
    COLOR_BLUE_LIGHT = 'F0F6FC'    # 部門元・候補2-3（薄い青色）
    COLOR_WHITE = 'FFFFFF'   # デフォルト（白）
    COLOR_HEADER_STREAMED = 'E2EFDA'  # STREAMEDの元の列（薄い緑色）
    COLOR_HEADER_SYSTEM = 'FCE4D6'    # システム追加列（薄いオレンジ色）
    COLOR_HEADER_GRAY = 'D9D9D9'      # その他のヘッダー（グレー）
    COLOR_MODIFIED = 'E7E6F7'         # 変更された列（薄い紫色）
    COLOR_SYSTEM_ADDED = 'DEEBF7'     # システムで追加/コピーされた列（薄い青色）

    def __init__(self):
        self.wb = None
        self.ws = None

    def write_to_excel(self, df, output_path, sheet_name='Sheet1'):
        """
        DataFrameをExcelに出力し、色分けを行う

        Args:
            df: 出力するDataFrame
            output_path: 出力先ファイルパス
            sheet_name: シート名

        Returns:
            str: 出力先ファイルパス
        """
        # 新しいワークブックを作成
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name

        # DataFrameをワークシートに書き込み
        columns = list(df.columns)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = self.ws.cell(row=r_idx, column=c_idx, value=value)

                # ヘッダー行のスタイル
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    # 列の種類に応じて色分け
                    col_name = columns[c_idx - 1] if c_idx <= len(columns) else ''
                    header_color = self._get_header_color(col_name)
                    cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
                # データ行で金額列の場合
                elif r_idx > 1 and c_idx <= len(columns):
                    col_name = columns[c_idx - 1]
                    if '金額' in col_name:
                        # 数値の場合、三桁カンマと右揃えを適用
                        if isinstance(value, (int, float)) and pd.notna(value):
                            cell.number_format = '#,##0'
                            cell.alignment = Alignment(horizontal='right', vertical='center')

        # 行の高さを1.5倍に設定
        self._set_row_height(df)

        # 列幅を自動調整
        self._auto_fit_columns(df)

        # 色分けを適用
        self._apply_colors(df)

        # 罫線を適用
        self._apply_borders(df)

        # 説明シートを追加
        self._add_instruction_sheet()

        # ファイルを保存
        self.wb.save(output_path)

        return output_path

    def _get_header_color(self, col_name):
        """
        列名に応じてヘッダーの色を返す

        Args:
            col_name: 列名

        Returns:
            str: 色コード
        """
        # システム追加列
        system_columns = [
            'STREAMED元の取引先', 'STREAMED元の部門',
            'freee取引先名候補1', 'freee取引先名候補2', 'freee取引先名候補3',
            'freee部門候補1', 'freee部門候補2', 'freee部門候補3'
        ]

        if col_name in system_columns:
            return self.COLOR_HEADER_SYSTEM

        # フラグ列（非表示になる列）
        if col_name.startswith('_'):
            return self.COLOR_HEADER_GRAY

        # STREAMED元の列（それ以外）
        return self.COLOR_HEADER_STREAMED

    def _set_row_height(self, df):
        """
        行の高さを1.5倍に設定

        Args:
            df: DataFrame
        """
        # デフォルトの行の高さは15、1.5倍にする
        default_height = 15
        new_height = default_height * 1.5

        # すべての行に適用（ヘッダー含む）
        for row_idx in range(1, len(df) + 2):  # ヘッダー + データ行
            self.ws.row_dimensions[row_idx].height = new_height

    def _auto_fit_columns(self, df):
        """
        列幅を自動調整（日本語対応）

        Args:
            df: DataFrame
        """
        for idx, column in enumerate(df.columns, 1):
            max_width = self._calculate_text_width(str(column))

            # データの最大幅を取得
            for value in df[column]:
                if pd.notna(value):
                    text_width = self._calculate_text_width(str(value))
                    if text_width > max_width:
                        max_width = text_width

            # 最大幅を設定（最大60、最小10）
            adjusted_width = min(max(max_width + 2, 10), 60)
            self.ws.column_dimensions[self._get_column_letter(idx)].width = adjusted_width

    def _calculate_text_width(self, text):
        """
        テキストの表示幅を計算（日本語対応）

        日本語文字は英数字の約2倍の幅を取るため、それを考慮する

        Args:
            text: テキスト

        Returns:
            float: 表示幅
        """
        width = 0
        for char in text:
            # 日本語文字（ひらがな、カタカナ、漢字、全角記号）
            if ord(char) > 127:
                width += 2
            else:
                width += 1
        return width

    def _get_column_letter(self, col_idx):
        """
        列番号をアルファベットに変換

        Args:
            col_idx: 列番号（1始まり）

        Returns:
            str: 列アルファベット（例: 1→A, 27→AA）
        """
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _apply_colors(self, df):
        """
        色分けを適用

        ルール:
        - 完全一致: 緑
        - 不一致: 赤
        - 取引先候補列: 黄色
        - 部門候補列: 青色

        Args:
            df: DataFrame
        """
        columns = list(df.columns)

        # 各行に対して処理
        for row_idx in range(len(df)):
            excel_row = row_idx + 2  # Excelの行番号（ヘッダーが1行目）

            # 取引先のチェック
            if '_取引先完全一致' in columns:
                is_partner_match = df.at[row_idx, '_取引先完全一致']
                has_partner = df.at[row_idx, 'STREAMED元の取引先'] if 'STREAMED元の取引先' in columns else ''

                if pd.notna(has_partner) and has_partner != '':
                    color = self.COLOR_GREEN if is_partner_match else self.COLOR_RED
                    self._color_row(excel_row, columns, color, exclude_patterns=['候補', '_'])

            # 部門のチェック
            if '_部門完全一致' in columns:
                is_dept_match = df.at[row_idx, '_部門完全一致']
                has_dept = df.at[row_idx, 'STREAMED元の部門'] if 'STREAMED元の部門' in columns else ''

                if pd.notna(has_dept) and has_dept != '':
                    # すでに取引先で色が付いていない場合のみ部門の色を適用
                    if '_取引先完全一致' not in columns or pd.isna(df.at[row_idx, 'STREAMED元の取引先']) or df.at[row_idx, 'STREAMED元の取引先'] == '':
                        color = self.COLOR_GREEN if is_dept_match else self.COLOR_RED
                        self._color_row(excel_row, columns, color, exclude_patterns=['候補', '_'])

        # 候補列に色を付ける
        self._color_candidate_columns(df, columns)

        # フラグ列（_で始まる列）を非表示にする
        self._hide_flag_columns(columns)

    def _color_row(self, excel_row, columns, color, exclude_patterns=None):
        """
        行全体に色を付ける

        Args:
            excel_row: Excelの行番号
            columns: 列名リスト
            color: 色コード
            exclude_patterns: 除外するパターンのリスト（列名に含まれる文字列）
        """
        if exclude_patterns is None:
            exclude_patterns = []

        for col_idx, col_name in enumerate(columns, 1):
            # 除外パターンに一致する列はスキップ
            skip = False
            for pattern in exclude_patterns:
                if pattern in col_name:
                    skip = True
                    break

            if not skip:
                cell = self.ws.cell(row=excel_row, column=col_idx)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    def _hide_flag_columns(self, columns):
        """
        フラグ列（_で始まる列）を非表示にする

        Args:
            columns: 列名リスト
        """
        for col_idx, col_name in enumerate(columns, 1):
            if col_name.startswith('_'):
                col_letter = self._get_column_letter(col_idx)
                self.ws.column_dimensions[col_letter].hidden = True

    def _color_candidate_columns(self, df, columns):
        """
        候補列に色を付ける

        Args:
            df: DataFrame
            columns: 列名リスト
        """
        for col_idx, col_name in enumerate(columns, 1):
            # STREAMED元の取引先列: 薄い黄色
            if col_name == 'STREAMED元の取引先':
                self._color_column(col_idx, len(df), self.COLOR_YELLOW_LIGHT)

            # freee取引先名候補1: 濃い黄色
            elif col_name == 'freee取引先名候補1':
                self._color_column(col_idx, len(df), self.COLOR_YELLOW)

            # freee取引先名候補2-3: 薄い黄色
            elif col_name in ['freee取引先名候補2', 'freee取引先名候補3']:
                self._color_column(col_idx, len(df), self.COLOR_YELLOW_LIGHT)

            # STREAMED元の部門列: 薄い青色
            elif col_name == 'STREAMED元の部門':
                self._color_column(col_idx, len(df), self.COLOR_BLUE_LIGHT)

            # freee部門候補1: 濃い青色
            elif col_name == 'freee部門候補1':
                self._color_column(col_idx, len(df), self.COLOR_BLUE)

            # freee部門候補2-3: 薄い青色
            elif col_name in ['freee部門候補2', 'freee部門候補3']:
                self._color_column(col_idx, len(df), self.COLOR_BLUE_LIGHT)

    def _color_column(self, col_idx, row_count, color):
        """
        特定の列に色を付ける

        Args:
            col_idx: 列番号（1始まり）
            row_count: 行数
            color: 色コード
        """
        for row_idx in range(2, row_count + 2):  # ヘッダーを除く
            cell = self.ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    def _apply_borders(self, df):
        """
        罫線を適用

        Args:
            df: DataFrame
        """
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        columns = list(df.columns)

        for row_idx, row in enumerate(self.ws.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                cell.border = thin_border

                # 金額列の場合は右揃えを維持、それ以外は左揃え
                if row_idx > 1 and col_idx <= len(columns):
                    col_name = columns[col_idx - 1]
                    if '金額' in col_name:
                        # 金額列は右揃えを維持
                        if cell.alignment is None or cell.alignment.horizontal != 'right':
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        # それ以外は左揃え
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    # ヘッダー行は左揃え
                    cell.alignment = Alignment(horizontal='left', vertical='center')

    def _add_instruction_sheet(self):
        """
        使い方説明シートを追加
        """
        # 新しいシートを作成（最後のシートとして）
        ws_instruction = self.wb.create_sheet("使い方")

        # 説明内容
        instructions = [
            ["STREAMED→freee会計インポート用csv修正アプリ 使い方"],
            [""],
            ["■ このファイルの確認方法"],
            [""],
            ["1. 行の色分け"],
            ["　緑色の行　→　freeeと取引先名/部門名が完全一致しています。チェック不要です。"],
            ["　赤色の行　→　freeeに該当する取引先名/部門名が見つかりませんでした。必ず確認してください。"],
            [""],
            ["2. 列の色分け"],
            ["　黄色の列　→　freee取引先名候補（濃い黄色が候補1、薄い黄色が候補2-3とSTREAMED元）"],
            ["　青色の列　→　freee部門候補（濃い青色が候補1、薄い青色が候補2-3とSTREAMED元）"],
            [""],
            ["3. 候補の選択方法"],
            ["　①「freee取引先名候補1」「freee部門候補1」の列を確認"],
            ["　②正しい候補が表示されていれば、そのままにする"],
            ["　③候補2や候補3が正しい場合は、候補1の列にコピー&ペースト"],
            ["　④候補がすべて間違っている場合は、手動で正しい名称を候補1に入力"],
            [""],
            ["■ ステージ2での処理"],
            [""],
            ["このExcelファイルを確認・修正後、ステージ2にアップロードすると："],
            [""],
            ["1. 「freee取引先名候補1」に値が入っている場合のみ、「貸方取引先」「借方取引先」に自動コピーされます"],
            ["2. 「freee部門候補1」に値が入っている場合のみ、「借方部門」「貸方部門」に自動コピーされます"],
            ["3. 複合仕訳の場合、同じ伝票番号内で取引先名と部門が統一されます"],
            ["4. freeeインポート用CSVが生成されます"],
            [""],
            ["■ 注意事項"],
            [""],
            ["・赤色の行は必ず内容を確認し、正しい取引先名/部門名を入力してください"],
            ["・緑色の行は確認不要ですが、念のため目視確認を推奨します"],
            ["・候補1に文字が入っている行のみ、ステージ2で自動的にコピーされます"],
            ["・候補1が空欄の行は、ステージ2でそのまま（空欄のまま）になります"],
        ]

        # データを書き込み
        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_instruction.cell(row=row_idx, column=col_idx, value=value)

                # タイトル行のスタイル
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    cell.font = Font(bold=True, size=14, color='FFFFFF')
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                # 見出し行のスタイル（■で始まる行）
                elif value and str(value).startswith('■'):
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                # 番号付きリスト
                elif value and (str(value).startswith('1.') or str(value).startswith('2.') or str(value).startswith('3.') or str(value).startswith('4.')):
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                # 通常の行
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 列幅を調整
        ws_instruction.column_dimensions['A'].width = 100

        # 行の高さを調整
        for row_idx in range(1, len(instructions) + 1):
            ws_instruction.row_dimensions[row_idx].height = 20

    def write_stage2_excel(self, original_df, processed_df, output_path):
        """
        ステージ2用のExcelを出力（2シート構成で変更点を強調）

        Args:
            original_df: 編集前のDataFrame（ステージ1の出力）
            processed_df: 編集後のDataFrame（ステージ2の出力）
            output_path: 出力先ファイルパス

        Returns:
            str: 出力先ファイルパス
        """
        # 新しいワークブックを作成
        self.wb = Workbook()

        # シート1: 編集前（ステージ1と同じスタイル）
        self.ws = self.wb.active
        self.ws.title = "編集前"
        self._write_sheet_data(original_df, is_stage2=False)

        # シート2: 編集後（変更点を強調）
        ws_after = self.wb.create_sheet("編集後（freeeインポート用）")
        self.ws = ws_after
        self._write_sheet_data(processed_df, is_stage2=True)

        # ファイルを保存
        self.wb.save(output_path)

        return output_path

    def _write_sheet_data(self, df, is_stage2=False):
        """
        シートにデータを書き込み、スタイルを適用

        Args:
            df: DataFrame
            is_stage2: ステージ2（編集後）かどうか
        """
        columns = list(df.columns)

        # DataFrameをワークシートに書き込み
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = self.ws.cell(row=r_idx, column=c_idx, value=value)

                # ヘッダー行のスタイル
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    col_name = columns[c_idx - 1] if c_idx <= len(columns) else ''

                    if is_stage2:
                        # ステージ2: 変更された列を強調
                        header_color = self._get_stage2_header_color(col_name)
                    else:
                        # ステージ1: 通常の色分け
                        header_color = self._get_header_color(col_name)

                    cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')

                # データ行で金額列の場合
                elif r_idx > 1 and c_idx <= len(columns):
                    col_name = columns[c_idx - 1]
                    if '金額' in col_name:
                        # 数値の場合、三桁カンマと右揃えを適用
                        if isinstance(value, (int, float)) and pd.notna(value):
                            cell.number_format = '#,##0'
                            cell.alignment = Alignment(horizontal='right', vertical='center')

        # 行の高さを1.5倍に設定
        self._set_row_height(df)

        # 列幅を自動調整
        self._auto_fit_columns(df)

        # 罫線を適用
        self._apply_borders(df)

    def _get_stage2_header_color(self, col_name):
        """
        ステージ2の列名に応じてヘッダーの色を返す（変更点を強調）

        Args:
            col_name: 列名

        Returns:
            str: 色コード
        """
        # システムで生成/変更された列
        if col_name == '伝票番号':
            return self.COLOR_MODIFIED  # 薄い紫色

        # 候補1からコピーされた列
        if col_name in ['借方取引先', '貸方取引先', '借方部門', '貸方部門']:
            return self.COLOR_SYSTEM_ADDED  # 薄い青色

        # それ以外はSTREAMED元の列
        return self.COLOR_HEADER_STREAMED
